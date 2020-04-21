from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait
import time
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font
import os.path

#Using Chrome driver
driver = webdriver.Chrome(executable_path="c:\selenium_drivers\chromedriver.exe")
driver.maximize_window()

#Find user location
driver.get("https://www.ip2location.com/")
city = driver.find_element_by_id('city-name').text
#Scroll until city name
scroll_till = driver.find_element_by_id("city-name")
driver.execute_script("arguments[0].scrollIntoView();", scroll_till)

#Wait until the city name element is loaded
while city == "Loading...":
    time.sleep(1)
    city = driver.find_element_by_id('city-name').text

#Googling weather for the city
driver.get("http://google.com")

#Find the Search field, type in city name and weather and press Enter
driver.find_element_by_name("q").send_keys("weather ", city, Keys.RETURN)

#Get the temperature & humidity
temp = driver.find_element_by_id("wob_tm").text
humid = driver.find_element_by_id("wob_hm").text

#See if temperature is in F or C
celsius_button = driver.find_element_by_css_selector("span[aria-label='°Celsius']")
if celsius_button.is_enabled() == True:
    temp_scale = "°C"
else:
    temp_scale = "°F"

#Check Date and Time
driver.get("https://www.worldtimeserver.com/")

#Wait until the time element is loaded
wait=WebDriverWait(driver,30)
element=wait.until(EC.visibility_of_element_located((By.ID,"theTime")))

date = driver.find_element_by_id("theDate").text
time = driver.find_element_by_id("theTime").text
timezone = driver.find_element_by_id("theTimeZone").text

#Print to console log
print("The temperature in",city,"is",temp,temp_scale,"and humidity is",humid)

driver.close()

#Open (or Create) the Excel file to write the weather data to
file = 'nice_weather_log.xlsx'

if not os.path.exists(file):
    print("Creating nice_weather_log.xlsx", end=" ")
    weather_log = Workbook()
    sheet = weather_log.active
    sheet.title = "Weather Log"
    sheet.cell(1, 1, "Date")
    sheet.cell(1, 2, "Time")
    sheet.cell(1, 3, "Temperature")
    sheet.cell(1, 4, "Humidity")
    sheet.cell(1, 5, "Time Zone")
 #Set columns width
    sheet.column_dimensions['A'].width = 21
    sheet.column_dimensions['B'].width = 12
    sheet.column_dimensions['C'].width = 14
    sheet.column_dimensions['D'].width = 12
    sheet.column_dimensions['E'].width = 30
 #Make first row bold
    for cell in sheet["1:1"]:
        cell.font = Font(bold=True)
    weather_log.save(filename = file)
else:
    print("Updating nice_weather_log.xlsx...", end=" ")

workbook = openpyxl.load_workbook(file)
sheet = workbook.active

#See how many rows and columns are filled already
rows = sheet.max_row
columns = sheet.max_column

#Write date, time, temperature, humidity, and time zone data
temp = temp+temp_scale
sheet.cell(rows+1,1,date)
sheet.cell(rows+1,2,time)
sheet.cell(rows+1,3,temp)
sheet.cell(rows+1,4,humid)
sheet.cell(rows+1,5,timezone)

workbook.save(file)
print("Work complete.")